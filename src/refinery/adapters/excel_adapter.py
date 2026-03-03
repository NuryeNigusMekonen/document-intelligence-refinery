from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..models import ExtractedDocument, ExtractedPage, ProvenanceRef, TableObject, TextBlock


class ExcelAdapter:
    name = "excel_adapter"

    def extract(self, file_path: Path, doc_id: str) -> ExtractedDocument:
        wb = load_workbook(file_path, data_only=True)
        pages: list[ExtractedPage] = []
        page_num = 1
        for sheet in wb.worksheets:
            blocks: list[TextBlock] = []
            tables: list[TableObject] = []
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            if max_row == 0 or max_col == 0:
                pages.append(ExtractedPage(page_number=page_num, width=1000.0, height=1000.0, blocks=[], tables=[], figures=[]))
                page_num += 1
                continue

            matrix = [[sheet.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]
            non_empty_rows = [i for i, row in enumerate(matrix, start=1) if any(cell is not None and str(cell).strip() for cell in row)]
            if not non_empty_rows:
                pages.append(ExtractedPage(page_number=page_num, width=1000.0, height=1000.0, blocks=[], tables=[], figures=[]))
                page_num += 1
                continue

            start = min(non_empty_rows)
            end = max(non_empty_rows)
            table_rows = matrix[start - 1 : end]
            headers = [str(v).strip() if v is not None else "" for v in table_rows[0]]
            rows = [[str(v).strip() if v is not None else "" for v in r] for r in table_rows[1:]] if len(table_rows) > 1 else []
            cell_range = f"A{start}:{chr(64 + max_col)}{end}" if max_col <= 26 else f"A{start}:Z{end}"
            prov = ProvenanceRef(
                doc_name=file_path.name,
                ref_type="excel_cells",
                sheet_name=sheet.title,
                cell_range=cell_range,
                section_path=[sheet.title],
                content_hash="pending",
            )
            tables.append(
                TableObject(
                    bbox=(0.0, 0.0, 1000.0, 1000.0),
                    headers=headers,
                    rows=rows,
                    reading_order=0,
                    confidence=0.95,
                    provenance=prov,
                )
            )
            blocks.append(
                TextBlock(
                    text=f"[SHEET] {sheet.title}",
                    bbox=(0.0, 0.0, 1000.0, 20.0),
                    reading_order=0,
                    confidence=0.95,
                    provenance=prov,
                )
            )
            pages.append(ExtractedPage(page_number=page_num, width=1000.0, height=1000.0, blocks=blocks, tables=tables, figures=[]))
            page_num += 1

        return ExtractedDocument(doc_id=doc_id, doc_name=file_path.name, pages=pages)
