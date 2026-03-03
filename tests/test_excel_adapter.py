from pathlib import Path

import pytest

from refinery.adapters.excel_adapter import ExcelAdapter
from refinery.chunking import Chunker
from refinery.config import Settings
from refinery.storage import ArtifactStore


@pytest.mark.skipif(__import__("importlib").util.find_spec("openpyxl") is None, reason="openpyxl not installed")
def test_excel_adapter_table_and_provenance(tmp_path: Path):
    from openpyxl import Workbook

    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    wb.save(path)

    adapter = ExcelAdapter()
    extracted = adapter.extract(path, "doc_xlsx")
    assert extracted.pages
    assert extracted.pages[0].tables

    settings = Settings(workspace_root=tmp_path)
    store = ArtifactStore(settings)
    ldus = Chunker(settings, store).run(extracted)
    assert ldus
    assert any(ldu.chunk_type == "table" for ldu in ldus)
    assert any(ref.ref_type == "excel_cells" for ldu in ldus for ref in ldu.page_refs)
